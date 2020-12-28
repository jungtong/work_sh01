#-*-coding:utf-8

import os
import sys
import time
import shutil	# ZIP
from datetime import date, timedelta, datetime
from pprint import pprint 

from multiprocessing import Pool

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup		# pip3 install lxml
# from BeautifulSoup import BeautifulSoup, NavigableString

# import pickle
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from googleapiclient.discovery import build # pip3 install --upgrade google-api-python-client
from httplib2 import Http
from oauth2client import file, client, tools # pip3 install oauth2client
import requests # pip3 install requests


# reload(sys)  
# sys.setdefaultencoding('utf-8')

keywords = [
'BESTSHOP',
'LG교원',
'LG매장',
'LG멤버십',
'LG박람회',
'LG혼수',
'베스트라이프',
'베스트샵',
'엘지교원',
'엘지매장',
'엘지멤버십',
'LG전자 매장'
]

wheres = [
'블로그',
'카페',
'뉴스'
# '웹페이지'
]

pages = [
'1',
'2',
'3'
]

skipHrefs = [
'www.ppomppu.co.kr/zboard/view.php?id=sponsor',
'cafe.naver.com/joonggonara',
'lg-bestshop'
]

skipWordsTitles = [
'유플러스',
'알바',
'시급',
'렌탈',
'중고',
'아르바이트',
'채용',
'리퍼브',
'가전플래너'
]

skipWordsBodys = [
'유플러스',
'알바',
'시급',
'렌탈'
]

yesterday = date.today() - timedelta(1)
today = date.today()

DATE_STRING = yesterday.strftime('%y%m%d')

DIR_PATH = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = DIR_PATH + '/Output/' + DATE_STRING
OUTPUT_ZIP_PATH = DIR_PATH + '/Output'
OUTPUT_ZIP_FILE = DATE_STRING + '.zip'

##### TELEGRAM #####
## 서형봇 ##
TELEGRAM_TOKEN = "bot911043358:AAGTh-AdcbFkofmkQ0gLaXajOBoDMgbcCmQ"
TELEGRAM_CHAT_ID = "-1001388503761"
def sendTelegramMessage(pushMessage):
	print("🌈 TEST" + pushMessage)
	requests.post("https://api.telegram.org/" + TELEGRAM_TOKEN + "/sendMessage?chat_id=" + TELEGRAM_CHAT_ID + "&text=" + pushMessage)
###################


def getQueryUrl(query, where, page, yesterday, today):
	fd = yesterday.strftime('%Y%m%d')
	td = today.strftime('%Y%m%d')

	if where == "블로그":
		# https://search.naver.com/search.naver?query=베스트샵&where=post&start=11&date_option=8&date_from=20181124&date_to=20181125
		# fd = yesterday.strftime('%Y%m%d')
		# td = today.strftime('%Y%m%d')
		return "https://search.naver.com/search.naver?where=post&query={0}&start={1}&nso=so%3Ar%2Cp%3Afrom{2}to{3}%2Ca%3Aall".format(query, page, fd, td)
	elif where == "카페":
		# https://search.naver.com/search.naver?where=article&query=베스트샵&date_option=6&date_from=2018.11.24&date_to=2018.11.25
		# fd = yesterday.strftime('%Y.%m.%d')
		# td = today.strftime('%Y.%m.%d')
		return "https://search.naver.com/search.naver?where=article&query={0}&start={1}&nso=so%3Ar%2Cp%3Afrom{2}to{3}%2Ca%3Aall".format(query, page, fd, td)
	elif where == "뉴스":
		# https://search.naver.com/search.naver?where=news&query=베스트샵&pd=3&ds=2018.11.24&de=2018.11.25
		# fd = yesterday.strftime('%Y.%m.%d')
		# td = today.strftime('%Y.%m.%d')
		return "https://search.naver.com/search.naver?where=news&query={0}&start={1}&nso=so%3Ar%2Cp%3Afrom{2}to{3}%2Ca%3Aall".format(query, page, fd, td)
	elif where == "웹페이지":
		# https://search.naver.com/search.naver?where=webkr&query=베스트샵&nso=so%3Ar%2Ca%3Aall%2Cp%3Afrom20181120to20181123
		# fd = yesterday.strftime('%Y%m%d')
		# td = today.strftime('%Y%m%d')
		return "https://search.naver.com/search.naver?where=webkr&query={0}&start={1}&nso=so%3Ar%2Cp%3Afrom{2}to{3}%2Ca%3Aall".format(query, page, fd, td)

def convertStringForQuery(inputStr):
	return {
        '1': '1',
        '2': '11',
        '3': '21'
    }.get(inputStr, 'err') #default

def getHTMLFromUrl(url, where, fileName):
	chrome_options = Options()
	chrome_options.add_argument("--headless")
	chrome_options.add_argument("--window-size=1920,2000")
	chrome_options.add_argument("--hide-scrollbars")
	driver = webdriver.Chrome(executable_path=DIR_PATH+'/chromedriver', options=chrome_options)

	driver.get(url)
	html = driver.page_source
	time.sleep(3)

	# print(html)


	isEmpty = False
	if where == "뉴스":
		soup = BeautifulSoup(html, "lxml")
		if None == soup.find('ul', {'class':'list_news'}):
			path = OUTPUT_PATH + '/결과없음_' + fileName
			print("		screenShot: " + path)
			driver.save_screenshot(path)
			isEmpty = True

	if isEmpty == False:
		xpathString = "//div[@class='api_subject_bx']"
		div = driver.find_element_by_xpath(xpathString)
		path = OUTPUT_PATH + '/' + fileName
		print("		screenShot: " + path)
		div.screenshot(path)

	# print(div.text)
	
	driver.quit()
	return html

def implWorkWithItem(item):
	url = item['url']
	where = item['where']
	fileName = item['fileName']
	keyword = item['keyword']
	page = item['page']

	result = False
	while not result:
		try:
			html = getHTMLFromUrl(url, where, fileName)
			result = True
		except Exception as err:
			time.sleep(1)
			print(err)
			pass

	# print(html)

	crawledList = CrawlerHtml(html, where, item)

	subResult = []
	listOfHref = [row[0] for row in subResult]
	for crawledData in crawledList:
		if crawledData[0] not in listOfHref:
			crawledData.append(url)
			crawledData.append(page+'P')
			subResult.append(crawledData)
	newKey = keyword + '_' + where
	return {newKey : subResult}

def DoWork():
	yDateTime = yesterday.strftime('%y%m%d')
	tDateTime = today.strftime('%y%m%d')
	sendTelegramMessage("[LGBestShop] CAPTURE START : " + yDateTime + " ~ " + tDateTime + "\n")

	allResult = {}
	count = 0
	fullCount = len(keywords) * len(wheres) * len(pages)

	targetItems = []
	for keyword in keywords:
		for where in wheres:

			for page in pages:
				count = count + 1
				fileName = tDateTime + '_' + keyword + '_' + where + '_' + page + 'P' + '.png'
				fileName.replace('웹페이지', '웹사이트')

				url = getQueryUrl(keyword, where, convertStringForQuery(page), yesterday, today)
				targetInfo = {}
				targetInfo['url'] = url
				targetInfo['where'] = where
				targetInfo['fileName'] = fileName
				targetInfo['keyword'] = keyword
				targetInfo['page'] = page
				targetItems.append(targetInfo)

	pool = Pool(4)
	crawlData = pool.map(implWorkWithItem, targetItems)

	resultData = {}
	for rawData in crawlData:
		if rawData == None:
			continue

		key = list(rawData.keys())[0]
		value = rawData[key]

		if key in resultData.keys():
			oldValue = resultData[key]
			for item in value:
				oldValue.append(item)

			resultData[key] = oldValue
		else:
			resultData[key] = value
	return resultData

# def CrawlerHtml(html, where, item):
# 	resultList = []

# 	soup = BeautifulSoup(html, "lxml")
# 	# print(soup.prettify().encode('UTF-8'))

# 	if (where == "블로그"):
# 		for li in soup.find('ul', {'class':'type01'}).findAll('li'):
# 			link = li.find('a', {'class':'sh_blog_title _sp_each_url _sp_each_title'})
# 			linkBody = li.find('dd', {'class':'sh_blog_passage'})
# 			if linkBody == None:
# 				bodyText = ""
# 			else:
# 				bodyText = linkBody.text

# 			# print(link)

# 			linkString = link['href'].replace('?Redirect=Log&logNo=', '/')
# 			result = [linkString, link.text, bodyText]
# 			resultList.append(result)
			
# 			# print("======")
# 			# print("\t" + linkString)
# 			# print("\t" + link.text)
# 			# print("\t" + linkBody.text)
# 	elif (where == "뉴스"):
# 		for li in soup.find('ul', {'class':'type01'}).findAll('li'):
# 			dt = li.find('dt')
# 			if dt == None:
# 				continue

# 			link = dt.find('a')
# 			linkBody = li.findAll('dd')[-1]
# 			if linkBody == None:
# 				bodyText = ""
# 			else:
# 				bodyText = linkBody.text

# 			linkString = link['href'].replace('?Redirect=Log&logNo=', '/')
# 			result = [linkString, link.text, bodyText]
# 			resultList.append(result)
# 			# print("======")
# 			# print("\t" + linkString)
# 			# print("\t" + link.text)
# 			# print("\t" + linkBody.text)
# 	elif (where == "카페"):
# 		for li in soup.find('ul', {'class':'type01'}).findAll('li'):
# 			dt = li.find('dt')
# 			if dt == None:
# 				continue

# 			link = dt.find('a')
# 			linkBody = li.find('dd', {'class':'sh_cafe_passage'})
# 			if linkBody == None:
# 				bodyText = ""
# 			else:
# 				bodyText = linkBody.text
			
# 			linkString = link['href'].replace('?Redirect=Log&logNo=', '/')
# 			result = [linkString, link.text, bodyText]
# 			resultList.append(result)
# 			# print("======")
# 			# print("\t" + linkString)
# 			# print("\t" + link.text)
# 			# print("\t" + linkBody.text)
# 	elif (where == "웹페이지"):
# 		for li in soup.find('ul', {'class':'lst_total'}).findAll('li'):
# 			link = li.find('a', {'class':'thumb'})
# 			linkUrl = link["href"]

# 			link_title = li.find('a', {'class':'link_tit'})
# 			linkTitle = link_title.text

# 			body = li.find('a', {'class':'total_dsc'})
# 			bodyText = body.text

# 			result = [linkUrl, linkTitle, bodyText]
# 			resultList.append(result)

# 			# print("======")
# 			# print("\t" + linkString)
# 			# print("\t" + link.text)
# 			# print("\t" + linkBody.text)

# 	for r in resultList[:]:
# 		isNeedRemove = False
# 		for skipHref in skipHrefs:
# 			if skipHref in r[0]:
# 				isNeedRemove = True
# 				break
# 		if isNeedRemove == True:
# 			resultList.remove(r)
# 			continue

# 		for skipWordsTitle in skipWordsTitles:
# 			if skipWordsTitle in r[1]:
# 				isNeedRemove = True
# 				break
# 		if isNeedRemove == True:
# 			resultList.remove(r)
# 			continue

# 		for skipWordsBody in skipWordsBodys:
# 			if skipWordsBody in r[2]:
# 				isNeedRemove = True
# 				break
# 		if isNeedRemove == True:
# 			resultList.remove(r)
# 			continue

# 	return resultList

def CrawlerHtml(html, where, item):
	resultList = []

	soup = BeautifulSoup(html, "lxml")
	# print(soup.prettify().encode('UTF-8'))

	if where == "뉴스":
		aaaa = soup.find('ul', {'class':'list_news'})
		if aaaa == None:
			print("=======")
			print("list_news is None")
			pprint(item)
			return resultList

		for li in soup.find('ul', {'class':'list_news'}).findAll('li', {'class':'bx'}):
			link = li.find('a', {'class':'news_tit'})

			if link == None:
				link = li.find('a', {'class':'elss sub_tit'})
				if link == None:
					print("== NEWS link 2 ==")
					pprint(li)

			linkUrl = link["href"]

			# link_title = li.find('a', {'class':'link_tit'})
			linkTitle = link.text

			body = li.find('div', {'class':'news_dsc'})
			if body == None:
				body = li.find('a', {'class':'dsc_txt_wrap'})
				if body == None:
					print("== NEWS 3==")
					pprint(item)
					pprint(li)

			bodyText = body.text

			result = [linkUrl, linkTitle, bodyText]
			resultList.append(result)
	else:
		aab = soup.find('ul', {'class':'lst_total'})
		if aab == None:
			print("=======")
			print("list_total is None")
			pprint(item)
			return resultList

		for li in soup.find('ul', {'class':'lst_total'}).findAll('li', {'class':'bx'}):
			link = li.find('a', {'class':'total_tit'})
			if link == None:
				link = li.find('a', {'class':'elss sub_tit'})
				if link == None:
					print("== NOT NEWS link 2 ==")
					pprint(li)

			linkUrl = link["href"]

			# link_title = li.find('a', {'class':'link_tit'})
			linkTitle = link.text

			body = li.find('a', {'class':'total_dsc'})
			if body == None:
				print("== NOT NEWS ==")
				pprint(li)
			bodyText = body.text

			result = [linkUrl, linkTitle, bodyText]
			resultList.append(result)

	for r in resultList[:]:
		isNeedRemove = False
		for skipHref in skipHrefs:
			if skipHref in r[0]:
				isNeedRemove = True
				break
		if isNeedRemove == True:
			resultList.remove(r)
			continue

		for skipWordsTitle in skipWordsTitles:
			if skipWordsTitle in r[1]:
				isNeedRemove = True
				break
		if isNeedRemove == True:
			resultList.remove(r)
			continue

		for skipWordsBody in skipWordsBodys:
			if skipWordsBody in r[2]:
				isNeedRemove = True
				break
		if isNeedRemove == True:
			resultList.remove(r)
			continue

	return resultList

def writeToExcel(rawData, fileName):
	wb = Workbook()
	# wb = load_workbook(fileName)
	# ws = wb.create_sheet("DataSheet")
	ws = wb.active

	ws.cell(row = 1, column=1).value = "키워드"
	ws.cell(row = 1, column=1).alignment = Alignment(horizontal='center')
	ws.cell(row = 1, column=2).value = "채널"
	ws.cell(row = 1, column=2).alignment = Alignment(horizontal='center')
	ws.cell(row = 1, column=3).value = "Page"
	ws.cell(row = 1, column=3).alignment = Alignment(horizontal='center')
	ws.cell(row = 1, column=4).value = "제목"
	ws.cell(row = 1, column=4).alignment = Alignment(horizontal='center')
	ws.cell(row = 1, column=5).value = "URL"
	ws.cell(row = 1, column=5).alignment = Alignment(horizontal='center')
	ws.cell(row = 1, column=6).value = "내용"
	ws.cell(row = 1, column=6).alignment = Alignment(horizontal='center')
	ws.cell(row = 1, column=7).value = "네이버 URL"
	ws.cell(row = 1, column=7).alignment = Alignment(horizontal='center')


	targetRow = 2
	for key in sorted(rawData.keys()):
		objList = rawData[key]
		for obj in objList:
			ws.cell(row = targetRow, column=1).value = key.split('_')[0]   # KEY
			ws.cell(row = targetRow, column=2).value = key.split('_')[1]   # KEY
			ws.cell(row = targetRow, column=3).value = obj[4]   # Page
			ws.cell(row = targetRow, column=4).value = obj[1]   # obj Title
			ws.cell(row = targetRow, column=5).value = obj[0]   # obj Href
			ws.cell(row = targetRow, column=6).value = obj[2]   # obj Body
			ws.cell(row = targetRow, column=7).value = obj[3]   # URL
			targetRow = targetRow+1

	ws.column_dimensions[get_column_letter(1)].width = 11	   	# KEY
	ws.column_dimensions[get_column_letter(2)].width = 9		# KEY
	ws.column_dimensions[get_column_letter(3)].width = 6		# Page
	ws.column_dimensions[get_column_letter(4)].width = 70		# obj Title
	ws.column_dimensions[get_column_letter(5)].width = 70		# obj Href
	ws.column_dimensions[get_column_letter(6)].width = 250		# obj Body
	ws.column_dimensions[get_column_letter(7)].width = 70		# URL
	wb.save(fileName)
	return


def testWithSavedRawData():
	with open(OUTPUT_PATH + '/Pickle_' + yesterday.strftime('%y%m%d'), 'rb') as handle:
		rawData = pickle.load(handle)

	writeToExcel(rawData, OUTPUT_PATH + '/' + yesterday.strftime('%y%m%d') + '.xlsx')
	return

def createZip():
	shutil.make_archive(OUTPUT_PATH, 'zip', OUTPUT_PATH)
	return

def uploadToGoogleDrive(filePath, fileName):
	try :
		import argparse
		flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
	except ImportError:
		flags = None

	SCOPES = 'https://www.googleapis.com/auth/drive.file'
	store = file.Storage(DIR_PATH + '/storage.json')
	creds = store.get()

	if not creds or creds.invalid:
		# print("make new storage data file ")
		flow = client.flow_from_clientsecrets(DIR_PATH + '/client_secret_drive.json', SCOPES)
		creds = tools.run_flow(flow, store, flags) \
		if flags else tools.run(flow, store)

	DRIVE = build('drive', 'v3', http=creds.authorize(Http()))

	metadata = {'name': fileName,
				'mimeType': None,
				'parents' : ["1JlFoE0WLO6tZlBeBE1O8ChYwX_jAiFEP"] # Google Drive Upload Folder ID
				}
	fileFullPath = filePath+'/'+fileName
	# print(fileFullPath)
	res = DRIVE.files().create(body=metadata, media_body=fileFullPath, fields='webViewLink, id, webContentLink').execute()
	if res:
		sendMesage = "[LGBestShop][" + yesterday.strftime('%y%m%d') + "]" + res.get('webContentLink')
	else:
		sendMesage = "[LGBestShop][" + yesterday.strftime('%y%m%d') + "]Error"
	sendTelegramMessage(sendMesage)

def RepresentsInt(s):
	try: 
		int(s)
		return True
	except ValueError:
		return False

def r4unner_main(fromDateString=None, toDateString=None):
	if fromDateString == None:
		yDateTime = (date.today() - timedelta(1)).strftime('%y%m%d')
	else:
		yDateTime = fromDateString

	if toDateString == None:
		tDateTime = date.today().strftime('%y%m%d')
	else:
		tDateTime = toDateString

	if len(yDateTime) != 6:
		sendTelegramMessage("[LGBestShop] From date error 1 : " + yDateTime + "\n")
		return None
	if RepresentsInt(yDateTime) == False:
		sendTelegramMessage("[LGBestShop] From date error 2 : " + yDateTime + "\n")
		return None
	if len(tDateTime) != 6:
		sendTelegramMessage("[LGBestShop] To date error 1 : " + tDateTime + "\n")
		return  None
	if RepresentsInt(tDateTime) == False:
		sendTelegramMessage("[LGBestShop] To date error 2 : " + tDateTime + "\n")
		return  None

	try:
		global yesterday
		global today
		yesterday = datetime.strptime(yDateTime, "%y%m%d")
		today = datetime.strptime(tDateTime, "%y%m%d")
	except Exception as err:
		sendTelegramMessage("[LGBestShop] date error 3", yDateTime, tDateTime)
		return None

	global DATE_STRING
	DATE_STRING = yesterday.strftime('%y%m%d')
	global OUTPUT_PATH
	OUTPUT_PATH = DIR_PATH + '/Output/' + DATE_STRING
	global OUTPUT_ZIP_PATH
	OUTPUT_ZIP_PATH = DIR_PATH + '/Output'
	global OUTPUT_ZIP_FILE
	OUTPUT_ZIP_FILE = DATE_STRING + '.zip'

	if not os.path.exists(OUTPUT_PATH):
	    os.makedirs(OUTPUT_PATH)

	sendTelegramMessage("[LGBestShop] CAPTURE START : " + yDateTime + " ~ " + tDateTime + "\n")
	rawData = DoWork()
	if rawData == None:
		return

	# with open(OUTPUT_PATH + '/Pickle_' + yesterday.strftime('%y%m%d'), 'wb') as handle:
	# 	pickle.dump(rawData, handle, protocol=pickle.HIGHEST_PROTOCOL)

	writeToExcel(rawData, OUTPUT_PATH + '/' + yesterday.strftime('%y%m%d') + '.xlsx')

	sendTelegramMessage("[LGBestShop] CAPTURE COMPLETE")

	createZip()
	uploadResult = uploadToGoogleDrive(OUTPUT_ZIP_PATH, OUTPUT_ZIP_FILE)
	if uploadResult:
		sendMesage = "[LGBestShop][" + yesterday.strftime('%y%m%d') + "]" + uploadResult.get('webContentLink')
	else:
		sendMesage = "[LGBestShop][" + yesterday.strftime('%y%m%d') + "]Error"
	sendTelegramMessage(sendMesage)
	sendTelegramMessage("[LGBestShop]Fin")

def main():
	# chrome_options = Options()
	# # chrome_options.add_argument("--headless")
	# # chrome_options.add_argument("--window-size=1920,2000")
	# chrome_options.add_argument("--hide-scrollbars")
	# driver = webdriver.Chrome(executable_path=DIR_PATH+'/chromedriver', options=chrome_options)

	if not os.path.exists(OUTPUT_PATH):
	    os.makedirs(OUTPUT_PATH)

	rawData = DoWork()
	# driver.close()

	with open(OUTPUT_PATH + '/Pickle_' + yesterday.strftime('%y%m%d'), 'wb') as handle:
		pickle.dump(rawData, handle, protocol=pickle.HIGHEST_PROTOCOL)

	writeToExcel(rawData, OUTPUT_PATH + '/' + yesterday.strftime('%y%m%d') + '.xlsx')

	sendTelegramMessage("[LGBestShop] CAPTURE COMPLETE")

	createZip()
	uploadToGoogleDrive(OUTPUT_ZIP_PATH, OUTPUT_ZIP_FILE)
	sendTelegramMessage("[LGBestShop]Fin")

if __name__ == "__main__":
	# main()
	r4unner_main("201227", "201228")

	# url = "https://search.naver.com/search.naver?where=article&query=LG교원&start=21&nso=so%3Ar%2Cp%3Afrom20201106to20201107%2Ca%3Aall"
	# where = "블로그"
	# fileName = "00005Test.png"
	# html = getHTMLFromUrl(url, where, fileName)

	# aaa = CrawlerHtml(html, where, "a")
	# print(aaa)

	# main()
